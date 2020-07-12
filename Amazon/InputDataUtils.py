# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook("InputData.xlsx")
ws = wb.active
inputrow = ws.max_column
inputcolumn = ws.max_row
inputdata = {}
for i in range(1,inputrow+1):
    inputdata.update({ws.cell(row = i,column = 1).value:ws.cell(row = i,column = 2).value})

def getUserName():
    return inputdata['username']

def getPassword():
    return inputdata['password']
